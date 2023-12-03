import threading
import urllib.parse

from PySide6.QtCore import *
from PySide6.QtWidgets import *
from PySide6.QtGui import *

from ui.ui import Ui_MainWindow
from qt_material import apply_stylesheet

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoSuchWindowException

import time
import pyperclip
import requests
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from datetime import datetime
import os
from urllib.parse import urlparse, parse_qs, urlencode


# 메인 ui 클래스
class MainWindow(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.model: QStandardItemModel = None
        self.setupUi(self)
        self.logic = MainLogic()

        self.__init()

        self.progressBar.setVisible(False)

        # 로그인 관련 로직
        self.lineEdit_pw.setEchoMode(QLineEdit.EchoMode.Password)
        self.lineEdit_pw.returnPressed.connect(self.loginEvent)
        self.pushButton_login.clicked.connect(self.loginEvent)

        self.create_model_and_set_table_top_header()

        header = self.tableView_result.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)

        self.pushButton_search.clicked.connect(self.searchEvent)

        self.result_directory = 'crawling_result'
        os.makedirs(self.result_directory, exist_ok=True)
        self.pushButton_excelSave.clicked.connect(self.excelSaveEvent)

    def __init(self):
        self.pushButton_search.setDisabled(True)
        self.pushButton_excelSave.setDisabled(True)

    def create_model_and_set_table_top_header(self):
        table_top_header_labels = ['이웃명', '포스트 제목', '포스트 썸네일', '공감 개수', '댓글 개수', '포스트 url']
        self.model = QStandardItemModel(None, 0, len(table_top_header_labels))
        self.model.setHorizontalHeaderLabels(table_top_header_labels)

        self.tableView_result.setModel(self.model)

    def __create_messagebox_when_web_broswer_closed(self):
        QMessageBox.warning(
            self,
            '웹브라우저 종료 오류',
            '''
            로그인을 다시 시도해주세요. 
            이 프로그램을 실행하는 동안 웹브라우저를 끄지 마세요.
            ''',
            QMessageBox.StandardButton.Ok,
        )

    def __check_web_broswer_alive(self) -> bool:
        if self.logic.is_web_broswer_alived() is False:
            self.__create_messagebox_when_web_broswer_closed()
            self.__init()
            return False
        return True

    def __set_progressing_ui(self):
        self.progressBar.setVisible(True)
        self.groupBox_control.setDisabled(True)

    def __unset_progressing_ui(self):
        self.progressBar.setVisible(False)
        self.groupBox_control.setEnabled(True)

    class LoginThread(QThread):
        def __init__(self, parent):
            super().__init__(parent)
            self.main_window: MainWindow = self.parent()

        def run(self):
            input_id = self.main_window.lineEdit_id.text()
            input_pw = self.main_window.lineEdit_pw.text()

            if self.main_window.logic.login(input_id, input_pw) is False:
                QMessageBox.warning(
                    self.main_window,
                    '로그인 오류',
                    '정확한 아이디와 비밀번호를 적어주세요.',
                    QMessageBox.StandardButton.Ok,
                )
                return

            self.main_window.pushButton_search.setEnabled(True)

            default_page_data = self.main_window.logic.find_default_page_data()
            self.main_window.lineEdit_startPage.setText(str(default_page_data[0]))
            self.main_window.lineEdit_endPage.setText(str(default_page_data[1]))

            self.main_window.comboBox_neighborGroup.addItems(self.main_window.logic.find_neighbor_group_list())

            self.quit()

    def loginEvent(self):
        self.logic.create_chrome_web_browser()

        login_thread = self.LoginThread(self)

        login_thread.started.connect(self.__set_progressing_ui)
        login_thread.finished.connect(self.__unset_progressing_ui)
        login_thread.start()

    def load_image_from_url(self, url: str):
        image = QImage()
        image.loadFromData(requests.get(url).content)
        pixmap = QPixmap.fromImage(image)
        return pixmap

    class SearchThread(QThread):
        def __init__(self, parent):
            super().__init__(parent)
            self.main_window: MainWindow = self.parent()

        def run(self):
            self.main_window.create_model_and_set_table_top_header()

            # 페이지 설정
            start_page = int(self.main_window.lineEdit_startPage.text())
            end_page = int(self.main_window.lineEdit_endPage.text())

            if start_page > end_page:
                QMessageBox.warning(
                    self.main_window,
                    '경고',
                    '시작 페이지가 끝 페이지보다 큰 값입니다.',
                    QMessageBox.StandardButton.Ok
                )

            self.main_window.logic.close_floating_popup()

            # 이웃그룹 설정
            neighbor_group_id = self.main_window.logic.get_neighbor_group_id(
                self.main_window.comboBox_neighborGroup.currentIndex()
            )

            for target_page in range(start_page, end_page+1):
                data = self.main_window.logic.get_neighbor_post_data(target_page, neighbor_group_id)

                for row in range(len(data)):
                    if self.isInterruptionRequested():
                        return

                    row_data = []

                    for column in range(len(data[0])):
                        item = QStandardItem()

                        if column == 2:
                            # 썸네일 가져오는 로직
                            if data[row][column] is not None:
                                pixmap = self.main_window.load_image_from_url(str(data[row][column]))
                                pixmap = pixmap.scaledToWidth(100)
                                item.setData(pixmap, Qt.ItemDataRole.DecorationRole)
                                # item.setSizeHint(pixmap.size())
                        else:
                            item.setText(str(data[row][column]))

                        row_data.append(item)

                    self.main_window.model.appendRow(row_data)

                    self.main_window.pushButton_excelSave.setEnabled(True)

            self.quit()

    def __set_progressing_ui_search_version(self):
        self.progressBar.setVisible(True)
        self.groupBox_userAuth.setDisabled(True)
        self.groupBox_filter.setDisabled(True)
        self.groupBox_result.setDisabled(True)

    def __unset_progressing_ui_search_version(self):
        self.progressBar.setVisible(False)
        self.groupBox_userAuth.setEnabled(True)
        self.groupBox_filter.setEnabled(True)
        self.groupBox_result.setEnabled(True)

    def searchEvent(self):
        button_text = self.pushButton_search.text()
        if button_text == '검색 시작':
            self.pushButton_search.setText('검색 중지')

            if self.__check_web_broswer_alive() is False:
                return

            self.search_thread = self.SearchThread(self)

            self.search_thread.started.connect(self.__set_progressing_ui_search_version)
            self.search_thread.finished.connect(self.__unset_progressing_ui_search_version)
            self.search_thread.start()
        elif button_text == '검색 중지':
            self.search_thread.requestInterruption()
            self.search_thread.wait()

            self.pushButton_search.setText('검색 시작')

    class ExcelSaveThread(QThread):
        def __init__(self, parent):
            super().__init__(parent)
            self.main_window: MainWindow = self.parent()

        def run(self):
            workbook: Workbook = Workbook()
            sheet: Worksheet = workbook.active

            top_header_count = self.main_window.model.columnCount()
            for col in range(top_header_count):
                header_item: QStandardItem = self.main_window.model.horizontalHeaderItem(col)
                sheet.cell(
                    row=1,
                    column=col+1,
                    value=str(header_item.text())
                )

            for row in range(self.main_window.model.rowCount()):
                for col in range(self.main_window.model.columnCount()):
                    item = self.main_window.model.item(row, col)
                    if item is not None:
                        sheet.cell(
                            row=row+2,
                            column=col+1,
                            value=str(item.text())
                        )

            result_file_name = os.path.join(
                self.main_window.result_directory,
                datetime.now().strftime("%Y%m%d_%H_%M_%S")
            )
            workbook.save(filename=f'{result_file_name}.xlsx')

            self.quit()

    def excelSaveEvent(self):
        if self.__check_web_broswer_alive() is False:
            return

        excel_save_thread = self.ExcelSaveThread(self)

        excel_save_thread.started.connect(self.__set_progressing_ui)
        excel_save_thread.finished.connect(self.__unset_progressing_ui)
        excel_save_thread.start()

    def closeEvent(self, event: QCloseEvent):
        result = QMessageBox.question(
            self,
            "종료 확인",
            "종료하시겠습니까?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if result == QMessageBox.StandardButton.Yes:
            self.logic.window_close()
            event.accept()
        else:
            event.ignore()


# 메인 로직 클래스
class MainLogic:
    def __init__(self):
        self.web_browser: WebDriver = None
        self.actions: ActionChains = None

        # 네이버 블로그 메인 url 쿼리 파싱
        naver_blog_main_url_query = urlparse(self.__naver_blog_main_url()).query
        self.naver_blog_main_url_querys = parse_qs(naver_blog_main_url_query)

    def __naver_blog_main_url(self) -> str:
        return 'https://section.blog.naver.com/BlogHome.naver?'

    def create_chrome_web_browser(self):
        if self.is_web_broswer_alived() is False:
            self.web_browser = webdriver.Chrome()
            self.actions = ActionChains(self.web_browser)

    def is_web_broswer_alived(self) -> bool:
        try:
            print(self.web_browser.current_url)
            return True
        except Exception:
            self.web_browser = None
            return False

    def login(self, input_id: str, input_pw: str) -> bool:
        self.web_browser.get(self.__naver_blog_main_url())
        time.sleep(2)

        # 로그인 상태가 아니라면 네이버 로그인 화면으로 이동
        login_button = self.web_browser.find_element(By.CLASS_NAME, "login_button")
        if login_button is None:
            return True
        else:
            login_button.click()

        # 캡챠 우회 위해 복붙+딜레이 활용
        pyperclip.copy(input_id)
        self.web_browser.find_element(value="id").send_keys(Keys.CONTROL, 'v')
        time.sleep(2)

        pyperclip.copy(input_pw)
        self.web_browser.find_element(value="pw").send_keys(Keys.CONTROL, 'v')
        time.sleep(2)

        self.web_browser.find_element(value="log.login").click()
        time.sleep(2)
        if self.__naver_blog_main_url() in self.web_browser.current_url:
            return True
        else:
            return False

    def find_default_page_data(self) -> []:  # [start_page, end_page]
        page_navigation = self.web_browser.find_element(
            by=By.XPATH,
            value='//*[@id="content"]/section/div[3]/div',
        )
        if page_navigation is None:
            return [1, 1]

        pages = page_navigation.find_elements(
            by=By.TAG_NAME,
            value='span'
        )
        start_page = 1
        end_page = len(pages)

        return [start_page, end_page]

    def find_neighbor_group_list(self) -> []:
        self.web_browser.find_element(
            by=By.XPATH,
            value='//*[@id="content"]/section/div[1]/div/div/a/i',
        ).click()
        time.sleep(0.5)

        neighbor_group_elements_list = self.web_browser.find_element(
            by=By.XPATH,
            value='//*[@id="content"]/section/div[1]/div/div/div',
        ).find_elements(
            by=By.TAG_NAME,
            value='a',
        )
        neighbor_group_list = list(
            map(
                lambda x: x.text,
                neighbor_group_elements_list
            )
        )
        return neighbor_group_list

    def get_neighbor_group_id(self, neighbor_group_index: int) -> int:
        self.__set_neighbor_group(neighbor_group_index)
        current_url_query = urlparse(self.web_browser.current_url).query
        return int(parse_qs(current_url_query)['groupId'][0])

    def __set_neighbor_group(self, neighbor_group_index: int):
        # 이웃 그룹 선택 콤보박스 닫기
        self.web_browser.find_element(
            by=By.CSS_SELECTOR,
            value='h3.title_heading',
        ).click()

        self.web_browser.find_element(
            by=By.XPATH,
            value='//*[@id="content"]/section/div[1]/div/div/a/i',
        ).click()
        time.sleep(0.5)

        self.web_browser.find_element(
            by=By.CSS_SELECTOR,
            value=f'a.item._buddy_dropdown_{neighbor_group_index}'
        ).click()
        time.sleep(0.5)

    def __get_naver_blog_main_target_page_url(self, target_page: int, neighbor_group_id: int) -> str:
        self.naver_blog_main_url_querys['currentPage'] = [str(target_page)]
        self.naver_blog_main_url_querys['groupId'] = [str(neighbor_group_id)]
        return self.__naver_blog_main_url() + urlencode(self.naver_blog_main_url_querys, encoding='UTF-8', doseq=True)

    def get_neighbor_post_data(self, target_page: int, neighbor_group_id: int) -> []:
        result_list = []

        self.web_browser.get(self.__get_naver_blog_main_target_page_url(target_page, neighbor_group_id))
        self.close_floating_popup()
        time.sleep(1)

        posts = None
        try:
            posts = self.web_browser.find_element(
                by=By.CSS_SELECTOR,
                value='div.list_post_article.list_post_article_comments'
            ).find_elements(
                by=By.CLASS_NAME,
                value='item_inner',
            )
        except NoSuchElementException:
            return result_list

        for post in posts:
            post_neighbor_name = post.find_element(
                by=By.CLASS_NAME,
                value='name_author',
            ).text

            post_title = post.find_element(
                by=By.CLASS_NAME,
                value='title_post',
            ).text

            post_thumnail_image_url = None
            try:
                post_thumnail_image_url = post.find_element(
                    by=By.CLASS_NAME,
                    value='img_post'
                ).get_attribute('bg-image')
            except NoSuchElementException:
                print("no thumnail image")

            post_heart_count = None
            try:
                post_heart_count = post.find_element(
                    by=By.CSS_SELECTOR,
                    value='em.u_cnt._count',
                ).text
            except NoSuchElementException:
                print("no post_heart")
                post_heart_count = 0

            post_comment_count = None
            try:
                post_comment_count = post.find_element(
                    by=By.CSS_SELECTOR,
                    value='span.reply',
                ).find_element(
                    by=By.TAG_NAME,
                    value='em',
                ).text
            except NoSuchElementException:
                print("no post_comment")
                post_comment_count = 0

            post_url = post.find_element(
                by=By.CLASS_NAME,
                value='desc_inner'
            ).get_attribute('ng-href')

            result_list.append([
                post_neighbor_name,
                post_title,
                post_thumnail_image_url,
                post_heart_count,
                post_comment_count,
                post_url,
            ])

        print(result_list)
        return result_list

    def close_floating_popup(self):
        try:
            popup = self.web_browser.find_element(
                by=By.ID,
                value='floatingda_home'
            )
            self.actions.move_to_element(popup).perform()
            time.sleep(0.3)

            self.web_browser.find_element(
                by=By.XPATH,
                value='//*[@id="floatingda_home"]/div/button'
            ).click()
        except NoSuchElementException:
            pass

    def window_close(self):
        if self.web_browser is not None:
            self.web_browser.close()


# Qt 애플리케이션 생성
app = QApplication()
window = MainWindow()
apply_stylesheet(app, theme='dark_blue.xml')

# 애플리케이션 실행
window.show()
app.exec()
